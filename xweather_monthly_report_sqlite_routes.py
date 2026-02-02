from flask import Blueprint, render_template, request, jsonify, current_app, g, make_response
import sqlite3
import csv
import io
from datetime import datetime
from openpyxl import load_workbook
import re

def to_float(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    # ambil angka pertama dari string: "10 km" -> 10, "4.5km" -> 4.5
    m = re.search(r"[-+]?\d*\.?\d+", s.replace(",", "."))
    if not m:
        return None
    try:
        return float(m.group(0))
    except:
        return None



xweather_report_bp = Blueprint("xweather_report", __name__)

# =========================
# SQLite helpers
# =========================
def _db_path():
    # kamu bisa set app.config["XWEATHER_DB_PATH"]
    return current_app.config.get("XWEATHER_DB_PATH", "xweather_reports.db")

def get_db():
    if "db" not in g:
        conn = sqlite3.connect(_db_path())
        conn.row_factory = sqlite3.Row
        g.db = conn
    return g.db

@xweather_report_bp.teardown_app_request
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()

def init_db():
    db = get_db()
    db.execute("""
    CREATE TABLE IF NOT EXISTS monthly_lightning_alerts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        report_month TEXT NOT NULL,           -- "YYYY-MM"
        severity TEXT NOT NULL,               -- "info"|"warning"|"alarm"
        asset_name TEXT NOT NULL,
        extent_km REAL,
        first_event_time TEXT,
        active_time TEXT,
        last_event_time TEXT,
        clear_time TEXT,
        duration_min INTEGER,
        strength_ka REAL,
        event_type TEXT,
        created_at TEXT NOT NULL
    )
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_month ON monthly_lightning_alerts(report_month)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_asset ON monthly_lightning_alerts(asset_name)")
    db.commit()

def _now_iso():
    return datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

def _norm_sev(s: str) -> str:
    s = (s or "").strip().lower()
    if s in ("alarm", "warning", "info"):
        return s
    # fallback: detect
    if "alarm" in s: return "alarm"
    if "warn" in s: return "warning"
    return "info"


# =========================
# Pages
# =========================
@xweather_report_bp.route("/xweather/monthly-report-editor")
def monthly_report_editor_page():
    # ensure DB exists
    init_db()
    return render_template("xweather_monthly_report_editor_sqlite.html")


# =========================
# API
# =========================
@xweather_report_bp.route("/api/xweather/monthly-report", methods=["GET"])
def api_list_monthly_report():
    init_db()
    month = (request.args.get("month") or "").strip()  # "YYYY-MM"
    if not month:
        return jsonify({"ok": False, "error": "missing month (YYYY-MM)"}), 400

    db = get_db()
    rows = db.execute("""
        SELECT * FROM monthly_lightning_alerts
        WHERE report_month = ?
        ORDER BY id DESC
    """, (month,)).fetchall()

    data = [dict(r) for r in rows]
    return jsonify({"ok": True, "month": month, "rows": data})


@xweather_report_bp.route("/api/xweather/monthly-report", methods=["POST"])
def api_create_monthly_report_row():
    init_db()
    payload = request.get_json(force=True, silent=True) or {}

    report_month = (payload.get("report_month") or "").strip()
    if not report_month:
        return jsonify({"ok": False, "error": "report_month required (YYYY-MM)"}), 400

    severity = _norm_sev(payload.get("severity"))
    asset_name = (payload.get("asset_name") or "").strip()
    if not asset_name:
        return jsonify({"ok": False, "error": "asset_name required"}), 400

    def _to_float(v):
        if v in ("", None): return None
        try: return float(v)
        except: return None

    def _to_int(v):
        if v in ("", None): return None
        try: return int(float(v))
        except: return None

    db = get_db()
    cur = db.execute("""
        INSERT INTO monthly_lightning_alerts(
            report_month, severity, asset_name, extent_km,
            first_event_time, active_time, last_event_time, clear_time,
            duration_min, strength_ka, event_type, created_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        report_month,
        severity,
        asset_name,
        _to_float(payload.get("extent_km")),
        (payload.get("first_event_time") or "").strip(),
        (payload.get("active_time") or "").strip(),
        (payload.get("last_event_time") or "").strip(),
        (payload.get("clear_time") or "").strip(),
        _to_int(payload.get("duration_min")),
        _to_float(payload.get("strength_ka")),
        (payload.get("event_type") or "").strip(),
        _now_iso()
    ))
    db.commit()

    new_id = cur.lastrowid
    row = db.execute("SELECT * FROM monthly_lightning_alerts WHERE id = ?", (new_id,)).fetchone()
    return jsonify({"ok": True, "row": dict(row)}), 201


@xweather_report_bp.route("/api/xweather/monthly-report/<int:row_id>", methods=["PUT"])
def api_update_monthly_report_row(row_id: int):
    init_db()
    payload = request.get_json(force=True, silent=True) or {}

    db = get_db()
    existing = db.execute("SELECT * FROM monthly_lightning_alerts WHERE id = ?", (row_id,)).fetchone()
    if not existing:
        return jsonify({"ok": False, "error": "row not found"}), 404

    def pick(key, default=None):
        return payload[key] if key in payload else default

    severity = _norm_sev(pick("severity", existing["severity"]))
    asset_name = (pick("asset_name", existing["asset_name"]) or "").strip()
    report_month = (pick("report_month", existing["report_month"]) or "").strip()

    if not asset_name:
        return jsonify({"ok": False, "error": "asset_name required"}), 400
    if not report_month:
        return jsonify({"ok": False, "error": "report_month required (YYYY-MM)"}), 400

    def _to_float(v, fallback):
        if v in ("", None): return None
        try: return float(v)
        except: return fallback

    def _to_int(v, fallback):
        if v in ("", None): return None
        try: return int(float(v))
        except: return fallback

    extent_km = _to_float(pick("extent_km", existing["extent_km"]), existing["extent_km"])
    duration_min = _to_int(pick("duration_min", existing["duration_min"]), existing["duration_min"])
    strength_ka = _to_float(pick("strength_ka", existing["strength_ka"]), existing["strength_ka"])

    db.execute("""
        UPDATE monthly_lightning_alerts SET
            report_month = ?,
            severity = ?,
            asset_name = ?,
            extent_km = ?,
            first_event_time = ?,
            active_time = ?,
            last_event_time = ?,
            clear_time = ?,
            duration_min = ?,
            strength_ka = ?,
            event_type = ?
        WHERE id = ?
    """, (
        report_month,
        severity,
        asset_name,
        extent_km,
        (pick("first_event_time", existing["first_event_time"]) or "").strip(),
        (pick("active_time", existing["active_time"]) or "").strip(),
        (pick("last_event_time", existing["last_event_time"]) or "").strip(),
        (pick("clear_time", existing["clear_time"]) or "").strip(),
        duration_min,
        strength_ka,
        (pick("event_type", existing["event_type"]) or "").strip(),
        row_id
    ))
    db.commit()

    row = db.execute("SELECT * FROM monthly_lightning_alerts WHERE id = ?", (row_id,)).fetchone()
    return jsonify({"ok": True, "row": dict(row)})


@xweather_report_bp.route("/api/xweather/monthly-report/<int:row_id>", methods=["DELETE"])
def api_delete_monthly_report_row(row_id: int):
    init_db()
    db = get_db()
    cur = db.execute("DELETE FROM monthly_lightning_alerts WHERE id = ?", (row_id,))
    db.commit()
    if cur.rowcount == 0:
        return jsonify({"ok": False, "error": "row not found"}), 404
    return jsonify({"ok": True, "deleted_id": row_id})


# =========================
# CSV export (by month)
# =========================
@xweather_report_bp.route("/xweather/monthly-report.csv", methods=["GET"])
def monthly_report_csv():
    init_db()
    month = (request.args.get("month") or "").strip()
    if not month:
        return "missing month (YYYY-MM)", 400

    db = get_db()
    rows = db.execute("""
        SELECT severity, asset_name, extent_km, first_event_time, active_time,
               last_event_time, clear_time, duration_min, strength_ka, event_type
        FROM monthly_lightning_alerts
        WHERE report_month = ?
        ORDER BY id DESC
    """, (month,)).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Severity","Asset name","Extent (km)","First event time","Active time",
        "Last event time","Clear time","Duration (min)","Strength (kA)","Type"
    ])
    for r in rows:
        writer.writerow([
            r["severity"], r["asset_name"], r["extent_km"], r["first_event_time"], r["active_time"],
            r["last_event_time"], r["clear_time"], r["duration_min"], r["strength_ka"], r["event_type"]
        ])

    resp = make_response(output.getvalue())
    resp.headers["Content-Type"] = "text/csv"
    resp.headers["Content-Disposition"] = f"attachment; filename=Xweather_Monthly_Report_{month}.csv"
    return resp

@xweather_report_bp.route("/api/xweather/monthly-report/assets", methods=["GET"])
def api_monthly_report_assets():
    init_db()
    month = (request.args.get("month") or "").strip()
    if not month:
        return jsonify({"ok": False, "error": "missing month (YYYY-MM)"}), 400

    db = get_db()
    rows = db.execute("""
        SELECT DISTINCT asset_name
        FROM monthly_lightning_alerts
        WHERE report_month = ?
        ORDER BY asset_name ASC
    """, (month,)).fetchall()

    assets = [r["asset_name"] for r in rows]
    return jsonify({"ok": True, "month": month, "assets": assets})

@xweather_report_bp.route("/xweather/monthly-report")
def xweather_monthly_report_viewer():
    init_db()
    return render_template("xweather_monthly_report_sqlite.html")


@xweather_report_bp.route("/api/xweather/monthly-report/import-csv", methods=["POST"])
def api_import_monthly_report_csv():
    init_db()

    f = request.files.get("file")
    report_month = (request.form.get("report_month") or "").strip()
    mode = (request.form.get("mode") or "append").strip().lower()

    if not f:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not report_month:
        return jsonify({"ok": False, "error": "report_month required (YYYY-MM)"}), 400
    if mode not in ("append", "replace"):
        return jsonify({"ok": False, "error": "mode must be append or replace"}), 400

    filename = f.filename.lower()
    rows_data = []

    # =========================
    # READ FILE
    # =========================
    if filename.endswith(".csv"):
        raw = f.read()
        try:
            text = raw.decode("utf-8-sig")
        except Exception:
            text = raw.decode("latin-1")

        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return jsonify({"ok": False, "error": "CSV has no header"}), 400

        rows_data = list(reader)

    elif filename.endswith(".xlsx"):
        wb = load_workbook(f, data_only=True)
    
        chosen_ws = None
        header_row_idx = 1
    
        # cari sheet + baris header yang berisi "Severity" & "Asset name"
        for ws in wb.worksheets:
            for r in range(1, 6):  # cek baris 1..5
                headers = [str(c.value).strip() if c.value else "" for c in ws[r]]
                hlow = [h.lower() for h in headers]
                if "severity" in hlow and "asset name" in hlow:
                    chosen_ws = ws
                    header_row_idx = r
                    break
            if chosen_ws:
                break
    
        if not chosen_ws:
            return jsonify({"ok": False, "error": "Excel: cannot find header row (Severity, Asset name)"}), 400
    
        headers = [str(c.value).strip() if c.value else "" for c in chosen_ws[header_row_idx]]
        rows_data = []
    
        for r in chosen_ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if all(v is None or str(v).strip()=="" for v in r):
                continue
            row = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
            rows_data.append(row)
    
    else:
        return jsonify({
            "ok": False,
            "error": "Unsupported file type. Use CSV, XLS, or XLSX"
        }), 400

    # =========================
    # HEADER MAPPING
    # =========================
    if not rows_data:
        return jsonify({"ok": False, "error": "No data rows found"}), 400

    header_map = {k.strip().lower(): k for k in rows_data[0].keys()}

    def h(*names):
        for n in names:
            if n.lower() in header_map:
                return header_map[n.lower()]
        return None

    H_SEV = h("Severity")
    H_ASSET = h("Asset name", "Asset")
    H_EXT = h("Extent (km)", "Extent")
    H_FIRST = h("First event time")
    H_ACTIVE = h("Active time")
    H_LAST = h("Last event time")
    H_CLEAR = h("Clear time")
    H_DUR = h("Duration (min)", "Duration")
    H_KA = h("Strength (kA)", "Strength")
    H_TYPE = h("Type", "Event type")

    if not H_SEV or not H_ASSET:
        return jsonify({
            "ok": False,
            "error": "CSV/Excel must contain at least: Severity, Asset name"
        }), 400

    def to_float(v):
        try: return float(v)
        except: return None

    def to_int(v):
        try: return int(float(v))
        except: return None

    db = get_db()

    if mode == "replace":
        db.execute("DELETE FROM monthly_lightning_alerts WHERE report_month = ?", (report_month,))
        db.commit()

    inserted, skipped = 0, 0

    try:
        db.execute("BEGIN")
        for row in rows_data:
            asset = str(row.get(H_ASSET) or "").strip()
            if not asset:
                skipped += 1
                continue

            db.execute("""
                INSERT INTO monthly_lightning_alerts(
                    report_month, severity, asset_name, extent_km,
                    first_event_time, active_time, last_event_time, clear_time,
                    duration_min, strength_ka, event_type, created_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                report_month,
                _norm_sev(row.get(H_SEV)),
                asset,
                to_float(row.get(H_EXT)),
                str(row.get(H_FIRST) or "").strip(),
                str(row.get(H_ACTIVE) or "").strip(),
                str(row.get(H_LAST) or "").strip(),
                str(row.get(H_CLEAR) or "").strip(),
                to_int(row.get(H_DUR)),
                to_float(row.get(H_KA)),
                str(row.get(H_TYPE) or "").strip(),
                _now_iso()
            ))
            inserted += 1

        db.commit()
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": f"import failed: {e}"}), 500

    return jsonify({
        "ok": True,
        "mode": mode,
        "report_month": report_month,
        "inserted": inserted,
        "skipped": skipped
    })

import re
from datetime import datetime
from collections import Counter

def _parse_hour_utc(s: str):
    """
    Extract hour (0-23) from first_event_time string.
    Supports: "05.01.2026 06:13", "2026-01-05 06:13", "2026-01-05T06:13:00Z", etc.
    """
    if not s:
        return None
    s = str(s).strip()

    # try common datetime parsing first
    fmts = [
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S",
    ]
    for f in fmts:
        try:
            return datetime.strptime(s, f).hour
        except Exception:
            pass

    # fallback: regex take first hh:mm
    m = re.search(r"\b([01]?\d|2[0-3])[:.]\d{2}\b", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    return None

def _percentile(sorted_vals, p):
    # p in [0..100]
    if not sorted_vals:
        return None
    if len(sorted_vals) == 1:
        return float(sorted_vals[0])
    k = (len(sorted_vals) - 1) * (p / 100.0)
    f = int(k)
    c = min(f + 1, len(sorted_vals) - 1)
    if f == c:
        return float(sorted_vals[f])
    return float(sorted_vals[f] + (sorted_vals[c] - sorted_vals[f]) * (k - f))

@xweather_report_bp.route("/api/xweather/monthly-report/expert", methods=["GET"])
def api_monthly_report_expert():
    init_db()
    month = (request.args.get("month") or "").strip()
    if not month:
        return jsonify({"ok": False, "error": "missing month (YYYY-MM)"}), 400

    db = get_db()
    rows = db.execute("""
        SELECT severity, first_event_time, duration_min, strength_ka, event_type
        FROM monthly_lightning_alerts
        WHERE report_month = ?
    """, (month,)).fetchall()

    total = len(rows)
    if total == 0:
        return jsonify({
            "ok": True,
            "month": month,
            "metrics": [
                {"metric": "Total events", "value": 0},
                {"metric": "Period (UTC)", "value": f"{month}"},
            ]
        })

    # severity stats
    sev_counts = Counter([(r["severity"] or "").strip().lower() for r in rows])
    alarm = sev_counts.get("alarm", 0)
    warning = sev_counts.get("warning", 0)
    info = sev_counts.get("info", 0)

    # duration stats (minutes)
    durations = []
    for r in rows:
        v = r["duration_min"]
        if v is None:
            continue
        try:
            durations.append(float(v))
        except Exception:
            pass
    durations.sort()

    dmin = durations[0] if durations else None
    dmax = durations[-1] if durations else None
    dmed = _percentile(durations, 50) if durations else None
    dmean = (sum(durations) / len(durations)) if durations else None
    q25 = _percentile(durations, 25) if durations else None
    q75 = _percentile(durations, 75) if durations else None
    iqr = (q75 - q25) if (q25 is not None and q75 is not None) else None
    p90 = _percentile(durations, 90) if durations else None

    # hour histogram from first_event_time
    hours = []
    for r in rows:
        h = _parse_hour_utc(r["first_event_time"])
        if h is not None:
            hours.append(h)
    hour_counts = Counter(hours)
    most_hour = hour_counts.most_common(1)[0][0] if hour_counts else None

    # busiest 3-hour window
    busiest_window = None
    busiest_count = None
    if hour_counts:
        # normalize to 0..23
        counts = [hour_counts.get(h, 0) for h in range(24)]
        best_h, best_sum = 0, -1
        for h in range(24):
            s3 = counts[h] + counts[(h+1)%24] + counts[(h+2)%24]
            if s3 > best_sum:
                best_sum = s3
                best_h = h
        busiest_window = f"{best_h:02d}:00–{(best_h+3)%24:02d}:00"
        busiest_count = best_sum

    # event type stats
    types = []
    cg_minus = 0
    for r in rows:
        t = (r["event_type"] or "").strip()
        if not t:
            continue
        tn = t.lower().replace(" ", "")
        types.append(t)
        if "cg-" in tn:
            cg_minus += 1
    dominant_type = Counter(types).most_common(1)[0][0] if types else None
    cg_minus_share = (cg_minus / total * 100.0) if total else 0.0

    # strength stats (|kA|)
    strengths = []
    for r in rows:
        v = r["strength_ka"]
        if v is None:
            continue
        try:
            strengths.append(abs(float(v)))
        except Exception:
            pass
    
    avg_abs_ka = (sum(strengths) / len(strengths)) if strengths else None
    min_abs_ka = min(strengths) if strengths else None
    max_abs_ka = max(strengths) if strengths else None
    

    # period display (best-effort)
    period_str = month

    metrics = [
        {"metric": "Total events", "value": total},
        {"metric": "Period (UTC)", "value": period_str},
        {"metric": "Fixed ring logic", "value": "Info=20 km; Warning=10 km; Alarm=4 km"},
        {"metric": "Alarm share", "value": f"{alarm} ({(alarm/total*100):.1f}%)"},
        {"metric": "Warning share", "value": f"{warning} ({(warning/total*100):.1f}%)"},
        {"metric": "Info share", "value": f"{info} ({(info/total*100):.1f}%)"},
    ]

    # duration block
    if durations:
        metrics += [
            {"metric": "Duration min / median / max (min)", "value": f"{dmin:.0f} / {dmed:.0f} / {dmax:.0f}"},
            {"metric": "Duration mean (min)", "value": f"{dmean:.2f}"},
            {"metric": "Duration IQR (Q75–Q25) (min)", "value": f"{iqr:.2f}" if iqr is not None else ""},
            {"metric": "Duration P90 (min)", "value": f"{p90:.2f}" if p90 is not None else ""},
        ]

    # hour block
    if most_hour is not None:
        metrics.append({"metric": "Most frequent hour (UTC)", "value": f"{most_hour}"})
    if busiest_window is not None:
        metrics.append({"metric": "Busiest 3-hour window (UTC)", "value": f"{busiest_window} ({busiest_count} events)"})

    # type & strength block
    if dominant_type:
        metrics.append({"metric": "Dominant type", "value": dominant_type})
    
    metrics.append({"metric": "CG- share", "value": f"{cg_minus_share:.1f}%"})
    
    if min_abs_ka is not None:
        metrics.append({"metric": "Minimum |peak current| (kA)", "value": f"{min_abs_ka:.2f}"})
    
    if avg_abs_ka is not None:
        metrics.append({"metric": "Average |peak current| (kA)", "value": f"{avg_abs_ka:.2f}"})
    
    if max_abs_ka is not None:
        metrics.append({"metric": "Maximum |peak current| (kA)", "value": f"{max_abs_ka:.2f}"})
    
    return jsonify({"ok": True, "month": month, "metrics": metrics})


