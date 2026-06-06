from flask import Blueprint, render_template, request, jsonify, g, make_response, Response
from werkzeug.exceptions import HTTPException
import os
import csv
import io
import re
import json
import threading
import queue
import time
import urllib.error
import urllib.request
from datetime import datetime
from collections import Counter
from collections import deque
import psycopg
from psycopg.rows import dict_row
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from flask_login import current_user, login_required
from routes.auth_utils import role_required
from dotenv import load_dotenv
load_dotenv()

xweather_report_bp = Blueprint("xweather_report", __name__)

DEFAULT_GOOGLE_DRIVE_FILE_ID = "19Pmm6dTxhI9QErH9rbcWKHX89hp7Q9JB"
DEFAULT_GOOGLE_DRIVE_FILE_URL = (
    "https://drive.google.com/file/d/"
    f"{DEFAULT_GOOGLE_DRIVE_FILE_ID}/view"
)


@xweather_report_bp.errorhandler(Exception)
def handle_xweather_api_error(error):
    if not request.path.startswith("/api/"):
        raise error

    status_code = 500
    message = str(error) or error.__class__.__name__
    if isinstance(error, HTTPException):
        status_code = error.code or status_code
        message = error.description

    return jsonify({"ok": False, "error": message}), status_code

# =========================================================
# PostgreSQL (Railway) - drop-in replacement for sqlite layer
# =========================================================
def get_db():
    if "db" not in g:
        dsn = os.getenv("DATABASE_URL")
        if not dsn:
            raise RuntimeError("DATABASE_URL not set. Add Railway PostgreSQL and ensure env is injected.")
        g.db = psycopg.connect(dsn, row_factory=dict_row, connect_timeout=5)
        # we manually commit/rollback on teardown
        g.db.autocommit = False
    return g.db


@xweather_report_bp.teardown_app_request
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        try:
            if exception:
                db.rollback()
            else:
                db.commit()
        except Exception:
            pass
        db.close()


def init_db():
    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
        CREATE TABLE IF NOT EXISTS monthly_lightning_alerts (
            id BIGSERIAL PRIMARY KEY,
            report_month TEXT NOT NULL,
            severity TEXT NOT NULL,
            asset_name TEXT NOT NULL,
            extent_km DOUBLE PRECISION,
            first_event_time TEXT,
            active_time TEXT,
            last_event_time TEXT,
            clear_time TEXT,
            duration_min INTEGER,
            strength_ka DOUBLE PRECISION,
            event_type TEXT,
            created_at TEXT NOT NULL
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS golf_locations (
            id BIGSERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            latitude DOUBLE PRECISION NOT NULL,
            longitude DOUBLE PRECISION NOT NULL,
            alarm_km DOUBLE PRECISION DEFAULT 4,
            warning_km DOUBLE PRECISION DEFAULT 10,
            info_km DOUBLE PRECISION DEFAULT 20,
            alarm_color TEXT DEFAULT '#EF4444',
            warning_color TEXT DEFAULT '#F97316',
            info_color TEXT DEFAULT '#FBBF24',
            -- === NEW: Rule #2 settings ===
            rule2_enabled BOOLEAN DEFAULT TRUE,
            rule2_min_strikes INTEGER DEFAULT 1,
            rule2_window_sec INTEGER DEFAULT 60,
            created_at TEXT NOT NULL
        );
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_mla_month ON monthly_lightning_alerts(report_month);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_mla_asset ON monthly_lightning_alerts(asset_name);")

        # untuk DB yang sudah ada tabelnya (safe migration)
        cur.execute("ALTER TABLE golf_locations ADD COLUMN IF NOT EXISTS rule2_enabled BOOLEAN DEFAULT TRUE;")
        cur.execute("ALTER TABLE golf_locations ADD COLUMN IF NOT EXISTS rule2_min_strikes INTEGER DEFAULT 1;")
        cur.execute("ALTER TABLE golf_locations ADD COLUMN IF NOT EXISTS rule2_window_sec INTEGER DEFAULT 60;")    
    db.commit()


_MONTH_RE = re.compile(r"^\d{4}-\d{2}$")


def _event_time_expr():
    return (
        "COALESCE(NULLIF(BTRIM(first_event_time), ''), NULLIF(BTRIM(active_time), ''), "
        "NULLIF(BTRIM(last_event_time), ''), NULLIF(BTRIM(clear_time), ''))"
    )


def _month_from_datetime_expr(column):
    dt = f"NULLIF(BTRIM({column}), '')"
    return f"""
        CASE
            WHEN {dt} ~ '^\\d{{4}}-\\d{{2}}' THEN substring({dt} from 1 for 7)
            WHEN {dt} ~ '^\\d{{2}}\\.\\d{{2}}\\.\\d{{4}}' THEN substring({dt} from 7 for 4) || '-' || substring({dt} from 4 for 2)
            ELSE NULL
        END
    """


def _event_month_expr():
    dt = _event_time_expr()
    return f"""
        CASE
            WHEN {dt} ~ '^\\d{{4}}-\\d{{2}}' THEN substring({dt} from 1 for 7)
            WHEN {dt} ~ '^\\d{{2}}\\.\\d{{2}}\\.\\d{{4}}' THEN substring({dt} from 7 for 4) || '-' || substring({dt} from 4 for 2)
            ELSE NULL
        END
    """


def _event_month_filter_expr():
    month_exprs = [
        _month_from_datetime_expr("first_event_time"),
        _month_from_datetime_expr("active_time"),
        _month_from_datetime_expr("last_event_time"),
        _month_from_datetime_expr("clear_time"),
    ]
    has_selected_month = " OR ".join(f"({expr}) = %s" for expr in month_exprs)
    no_other_month = " AND ".join(f"(({expr}) IS NULL OR ({expr}) = %s)" for expr in month_exprs)
    return f"(({has_selected_month}) AND ({no_other_month}))"


def _event_month_filter_params(month):
    return (month,) * 8


def _event_sort_expr():
    dt = _event_time_expr()
    return f"""
        CASE
            WHEN {dt} ~ '^\\d{{4}}-\\d{{2}}-\\d{{2}}' THEN replace(replace({dt}, 'T', ' '), 'Z', '')
            WHEN {dt} ~ '^\\d{{2}}\\.\\d{{2}}\\.\\d{{4}}' THEN substring({dt} from 7 for 4) || '-' || substring({dt} from 4 for 2) || '-' || substring({dt} from 1 for 2) || substring({dt} from 11)
            ELSE {dt}
        END
    """


def _month_arg():
    month = (request.args.get("month") or "").strip()
    if not _MONTH_RE.match(month):
        return None
    return month


def _sheet_name_from_month(report_month):
    return str(report_month or "").strip().replace("-", "")


def _extract_google_file_id(url):
    url = str(url or "").strip()
    patterns = [
        r"/file/d/([A-Za-z0-9_-]+)",
        r"/spreadsheets/d/([A-Za-z0-9_-]+)",
        r"[?&]id=([A-Za-z0-9_-]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return None


def _google_download_candidates(url):
    url = str(url or "").strip()
    file_id = _extract_google_file_id(url)
    candidates = []

    if file_id:
        candidates.append(
            f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
        )
        candidates.append(
            f"https://drive.google.com/uc?export=download&id={file_id}"
        )
    if url:
        candidates.append(url)

    deduped = []
    seen = set()
    for candidate in candidates:
        if candidate not in seen:
            deduped.append(candidate)
            seen.add(candidate)
    return deduped


def _download_google_workbook(url):
    errors = []
    for candidate in _google_download_candidates(url):
        try:
            req = urllib.request.Request(candidate, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=30) as res:
                raw = res.read()
                content_type = (res.headers.get("Content-Type") or "").lower()

            head = raw[:500].decode("utf-8", errors="ignore").lower()
            if "<html" in head or "<!doctype" in head:
                errors.append(f"{candidate}: returned HTML page")
                continue

            return raw, content_type, candidate
        except urllib.error.HTTPError as e:
            errors.append(f"{candidate}: HTTP {e.code}")
        except Exception as e:
            errors.append(f"{candidate}: {e}")

    raise RuntimeError("; ".join(errors) or "no download URL candidates")


def _worksheet_rows(ws, header_row_idx):
    headers = [str(c.value).strip() if c.value else "" for c in ws[header_row_idx]]
    rows_data = []
    for rr in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None or str(v).strip() == "" for v in rr):
            continue
        row = {headers[i]: (rr[i] if i < len(rr) else None) for i in range(len(headers))}
        rows_data.append(row)
    return rows_data


def _find_header_row(ws):
    for r in range(1, min(ws.max_row, 12) + 1):
        headers = [str(c.value).strip() if c.value else "" for c in ws[r]]
        hlow = [h.lower() for h in headers]
        if "severity" in hlow and "asset name" in hlow:
            return r
    return None


def _rows_from_workbook(file_obj, report_month=None, require_month_sheet=False):
    wb = load_workbook(file_obj, data_only=True)
    target_sheet = _sheet_name_from_month(report_month)

    chosen_ws = None
    if target_sheet and target_sheet in wb.sheetnames:
        chosen_ws = wb[target_sheet]
    elif require_month_sheet:
        return [], {
            "ok": False,
            "error": f"Sheet {target_sheet} tidak ditemukan. Nama sheet harus mengikuti bulan laporan (YYYYMM).",
            "available_sheets": wb.sheetnames,
            "required_sheet": target_sheet,
        }, 400

    if chosen_ws:
        header_row_idx = _find_header_row(chosen_ws)
        if not header_row_idx:
            return [], {
                "ok": False,
                "error": f"Sheet {chosen_ws.title}: cannot find header row (Severity, Asset name)",
                "selected_sheet": chosen_ws.title,
            }, 400
        return _worksheet_rows(chosen_ws, header_row_idx), {
            "selected_sheet": chosen_ws.title,
            "available_sheets": wb.sheetnames,
        }, 200

    for ws in wb.worksheets:
        header_row_idx = _find_header_row(ws)
        if header_row_idx:
            return _worksheet_rows(ws, header_row_idx), {
                "selected_sheet": ws.title,
                "available_sheets": wb.sheetnames,
            }, 200

    return [], {"ok": False, "error": "Excel: cannot find header row (Severity, Asset name)"}, 400


# =========================================================
# Helpers (keep same behavior)
# =========================================================
def _now_iso():
    return datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")


def _norm_sev(s: str) -> str:
    s = (s or "").strip().lower()
    if s in ("alarm", "warning", "info"):
        return s
    if "alarm" in s:
        return "alarm"
    if "warn" in s:
        return "warning"
    return "info"


def _to_float(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    # allow "10 km", "4.5km", "23,1"
    m = re.search(r"[-+]?\d*\.?\d+", s.replace(",", "."))
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _to_int(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _parse_hour_utc(s: str):
    """
    Extract hour 0..23 from common timestamp strings.
    Works with examples like: "05.01.2026 06:13", "2026-01-05 06:13", "2026-01-05T06:13:00Z".
    """
    if not s:
        return None
    s = str(s).strip()

    fmts = [
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S",
    ]
    for f in fmts:
        try:
            return datetime.strptime(s, f).hour
        except Exception:
            pass

    m = re.search(r"\b([01]?\d|2[0-3])[:.]\d{2}\b", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    return None


def _percentile(sorted_vals, p):
    if not sorted_vals:
        return None
    if len(sorted_vals) == 1:
        return float(sorted_vals[0])
    k = (len(sorted_vals) - 1) * (p / 100.0)
    f = int(k)
    c = min(f + 1, len(sorted_vals) - 1)
    if f == c:
        return float(sorted_vals[f])
    return float(sorted_vals[f] + (sorted_vals[c] - sorted_vals[f]) * (k - f))


def _import_monthly_rows(rows_data, report_month, mode):
    if not rows_data:
        return {"ok": False, "error": "No data rows found"}, 400

    header_map = {k.strip().lower(): k for k in rows_data[0].keys()}

    def h(*names):
        for n in names:
            key = n.strip().lower()
            if key in header_map:
                return header_map[key]
        return None

    H_SEV = h("Severity")
    H_ASSET = h("Asset name", "Asset")
    H_EXT = h("Extent (km)", "Extent")
    H_FIRST = h("First event time", "First")
    H_ACTIVE = h("Active time", "Active")
    H_LAST = h("Last event time", "Last")
    H_CLEAR = h("Clear time", "Clear")
    H_DUR = h("Duration (min)", "Duration")
    H_KA = h("Strength (kA)", "Strength")
    H_TYPE = h("Type", "Event type")

    if not H_SEV or not H_ASSET:
        return {"ok": False, "error": "Data must contain at least: Severity, Asset name"}, 400

    db = get_db()
    inserted = 0
    skipped = 0

    try:
        with db.cursor() as cur:
            if mode == "replace":
                cur.execute("DELETE FROM monthly_lightning_alerts WHERE report_month = %s", (report_month,))

            for row in rows_data:
                asset = str(row.get(H_ASSET) or "").strip()
                if not asset:
                    skipped += 1
                    continue

                cur.execute("""
                    INSERT INTO monthly_lightning_alerts(
                        report_month, severity, asset_name, extent_km,
                        first_event_time, active_time, last_event_time, clear_time,
                        duration_min, strength_ka, event_type, created_at
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    report_month,
                    _norm_sev(row.get(H_SEV)),
                    asset,
                    _to_float(row.get(H_EXT)),
                    str(row.get(H_FIRST) or "").strip(),
                    str(row.get(H_ACTIVE) or "").strip(),
                    str(row.get(H_LAST) or "").strip(),
                    str(row.get(H_CLEAR) or "").strip(),
                    _to_int(row.get(H_DUR)),
                    _to_float(row.get(H_KA)),
                    str(row.get(H_TYPE) or "").strip(),
                    _now_iso()
                ))
                inserted += 1

        db.commit()
    except Exception as e:
        db.rollback()
        return {"ok": False, "error": f"import failed: {e}"}, 500

    return {
        "ok": True,
        "mode": mode,
        "report_month": report_month,
        "inserted": inserted,
        "skipped": skipped
    }, 200


# =========================================================
# Pages (HTML stays the same)
# =========================================================
@xweather_report_bp.route("/xweather/monthly-report")
def xweather_monthly_report_viewer():
    return render_template("project/xweather_monthly_report_sqlite.html")


@xweather_report_bp.route("/xweather/monthly-report-editor")
@login_required
@role_required("admin")
def xweather_monthly_report_editor():
    init_db()
    return render_template("project/xweather_monthly_report_editor_sqlite.html")


# =========================================================
# API: List data (viewer/editor use this)
# GET /api/xweather/monthly-report?month=YYYY-MM
# =========================================================
@xweather_report_bp.route("/api/xweather/monthly-report", methods=["GET"])
def api_list_monthly_report():
    month = _month_arg()
    if not month:
        return jsonify({"ok": False, "error": "missing month (YYYY-MM)"}), 400

    init_db()
    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            SELECT *
            FROM monthly_lightning_alerts
            WHERE report_month = %s
            ORDER BY id DESC
        """, (month,))
        rows = cur.fetchall()

    return jsonify({"ok": True, "month": month, "rows": rows})


# =========================================================
# API: Assets dropdown
# GET /api/xweather/monthly-report/assets?month=YYYY-MM
# =========================================================
@xweather_report_bp.route("/api/xweather/monthly-report/assets", methods=["GET"])
def api_monthly_report_assets():
    month = _month_arg()
    if not month:
        return jsonify({"ok": False, "error": "missing month (YYYY-MM)"}), 400

    init_db()
    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT asset_name
            FROM monthly_lightning_alerts
            WHERE report_month = %s
            ORDER BY asset_name ASC
        """, (month,))
        rows = cur.fetchall()

    assets = [r["asset_name"] for r in rows]
    return jsonify({"ok": True, "month": month, "assets": assets})


# =========================================================
# API: Create row
# POST /api/xweather/monthly-report
# =========================================================
@xweather_report_bp.route("/api/xweather/monthly-report", methods=["POST"])
def api_create_monthly_report_row():
    init_db()
    payload = request.get_json(force=True, silent=True) or {}

    report_month = (payload.get("report_month") or "").strip()
    if not report_month:
        return jsonify({"ok": False, "error": "report_month required (YYYY-MM)"}), 400

    severity = _norm_sev(payload.get("severity"))
    asset_name = (payload.get("asset_name") or "").strip()
    if not asset_name:
        return jsonify({"ok": False, "error": "asset_name required"}), 400

    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            INSERT INTO monthly_lightning_alerts(
                report_month, severity, asset_name, extent_km,
                first_event_time, active_time, last_event_time, clear_time,
                duration_min, strength_ka, event_type, created_at
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING *
        """, (
            report_month,
            severity,
            asset_name,
            _to_float(payload.get("extent_km")),
            (payload.get("first_event_time") or "").strip(),
            (payload.get("active_time") or "").strip(),
            (payload.get("last_event_time") or "").strip(),
            (payload.get("clear_time") or "").strip(),
            _to_int(payload.get("duration_min")),
            _to_float(payload.get("strength_ka")),
            (payload.get("event_type") or "").strip(),
            _now_iso()
        ))
        row = cur.fetchone()

    db.commit()
    return jsonify({"ok": True, "row": row}), 201


# =========================================================
# API: Update row
# PUT /api/xweather/monthly-report/<id>
# =========================================================
@xweather_report_bp.route("/api/xweather/monthly-report/<int:row_id>", methods=["PUT"])
def api_update_monthly_report_row(row_id: int):
    init_db()
    payload = request.get_json(force=True, silent=True) or {}
    db = get_db()

    with db.cursor() as cur:
        cur.execute("SELECT * FROM monthly_lightning_alerts WHERE id = %s", (row_id,))
        existing = cur.fetchone()

    if not existing:
        return jsonify({"ok": False, "error": "row not found"}), 404

    report_month = (payload.get("report_month") if "report_month" in payload else existing["report_month"]) or ""
    report_month = str(report_month).strip()
    if not report_month:
        return jsonify({"ok": False, "error": "report_month required (YYYY-MM)"}), 400

    asset_name = (payload.get("asset_name") if "asset_name" in payload else existing["asset_name"]) or ""
    asset_name = str(asset_name).strip()
    if not asset_name:
        return jsonify({"ok": False, "error": "asset_name required"}), 400

    severity = _norm_sev(payload.get("severity", existing["severity"]))

    extent_km = _to_float(payload.get("extent_km", existing["extent_km"]))
    duration_min = _to_int(payload.get("duration_min", existing["duration_min"]))
    strength_ka = _to_float(payload.get("strength_ka", existing["strength_ka"]))

    first_event_time = (payload.get("first_event_time", existing["first_event_time"]) or "")
    active_time = (payload.get("active_time", existing["active_time"]) or "")
    last_event_time = (payload.get("last_event_time", existing["last_event_time"]) or "")
    clear_time = (payload.get("clear_time", existing["clear_time"]) or "")
    event_type = (payload.get("event_type", existing["event_type"]) or "")

    with db.cursor() as cur:
        cur.execute("""
            UPDATE monthly_lightning_alerts SET
                report_month = %s,
                severity = %s,
                asset_name = %s,
                extent_km = %s,
                first_event_time = %s,
                active_time = %s,
                last_event_time = %s,
                clear_time = %s,
                duration_min = %s,
                strength_ka = %s,
                event_type = %s
            WHERE id = %s
            RETURNING *
        """, (
            report_month,
            severity,
            asset_name,
            extent_km,
            str(first_event_time).strip(),
            str(active_time).strip(),
            str(last_event_time).strip(),
            str(clear_time).strip(),
            duration_min,
            strength_ka,
            str(event_type).strip(),
            row_id
        ))
        row = cur.fetchone()

    db.commit()
    return jsonify({"ok": True, "row": row})


# =========================================================
# API: Delete row
# DELETE /api/xweather/monthly-report/<id>
# =========================================================
@xweather_report_bp.route("/api/xweather/monthly-report/<int:row_id>", methods=["DELETE"])
def api_delete_monthly_report_row(row_id: int):
    init_db()
    db = get_db()
    with db.cursor() as cur:
        cur.execute("DELETE FROM monthly_lightning_alerts WHERE id = %s RETURNING id", (row_id,))
        deleted = cur.fetchone()

    db.commit()
    if not deleted:
        return jsonify({"ok": False, "error": "row not found"}), 404

    return jsonify({"ok": True, "deleted_id": row_id})


# =========================================================
# CSV export
# GET /xweather/monthly-report.csv?month=YYYY-MM
# =========================================================
@xweather_report_bp.route("/xweather/monthly-report.csv", methods=["GET"])
def monthly_report_csv():
    month = _month_arg()
    if not month:
        return "missing month (YYYY-MM)", 400

    init_db()
    db = get_db()
    with db.cursor() as cur:
        cur.execute(f"""
            SELECT severity, asset_name, extent_km, first_event_time, active_time,
                   last_event_time, clear_time, duration_min, strength_ka, event_type
            FROM monthly_lightning_alerts
            WHERE {_event_month_filter_expr()}
            ORDER BY {_event_sort_expr()} DESC NULLS LAST, id DESC
        """, _event_month_filter_params(month))
        rows = cur.fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Severity", "Asset name", "Extent (km)", "First event time", "Active time",
        "Last event time", "Clear time", "Duration (min)", "Strength (kA)", "Type"
    ])
    for r in rows:
        writer.writerow([
            r["severity"], r["asset_name"], r["extent_km"], r["first_event_time"], r["active_time"],
            r["last_event_time"], r["clear_time"], r["duration_min"], r["strength_ka"], r["event_type"]
        ])

    resp = make_response(output.getvalue())
    resp.headers["Content-Type"] = "text/csv"
    resp.headers["Content-Disposition"] = f"attachment; filename=Xweather_Monthly_Report_{month}.csv"
    return resp


# =========================================================
# Import CSV / XLSX
# POST /api/xweather/monthly-report/import-csv (multipart)
# =========================================================
@xweather_report_bp.route("/api/xweather/monthly-report/import-csv", methods=["POST"])
def api_import_monthly_report_csv():
    init_db()

    f = request.files.get("file")
    report_month = (request.form.get("report_month") or "").strip()
    mode = (request.form.get("mode") or "append").strip().lower()

    if not f:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not report_month:
        return jsonify({"ok": False, "error": "report_month required (YYYY-MM)"}), 400
    if mode not in ("append", "replace"):
        return jsonify({"ok": False, "error": "mode must be append or replace"}), 400

    filename = (f.filename or "").lower()
    rows_data = []

    # ---------- CSV ----------
    if filename.endswith(".csv"):
        raw = f.read()
        try:
            text = raw.decode("utf-8-sig")
        except Exception:
            text = raw.decode("latin-1")

        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return jsonify({"ok": False, "error": "CSV has no header"}), 400
        rows_data = list(reader)

    # ---------- XLSX ----------
    elif filename.endswith(".xlsx"):
        rows_data, meta, meta_status = _rows_from_workbook(
            f,
            report_month=report_month,
            require_month_sheet=False,
        )
        if meta_status != 200:
            return jsonify(meta), meta_status

    else:
        # keep behavior strict: only CSV/XLSX
        return jsonify({"ok": False, "error": "Unsupported file type. Use CSV or XLSX"}), 400

    result, status = _import_monthly_rows(rows_data, report_month, mode)
    if filename.endswith(".xlsx") and status == 200:
        result["selected_sheet"] = meta.get("selected_sheet")
    return jsonify(result), status


@xweather_report_bp.route("/api/xweather/monthly-report/sync-google-sheet", methods=["POST"])
def api_sync_monthly_report_google_sheet():
    payload = request.get_json(silent=True) or {}
    sync_token = os.getenv("XWEATHER_GOOGLE_SYNC_TOKEN", "").strip()
    supplied_token = (
        payload.get("token")
        or request.form.get("token")
        or request.args.get("token")
        or request.headers.get("X-Sync-Token")
        or ""
    )
    token_ok = bool(sync_token and supplied_token == sync_token)
    try:
        user_ok = bool(current_user.is_authenticated and getattr(current_user, "role", "") == "admin")
    except Exception:
        user_ok = False
    if not token_ok and not user_ok:
        return jsonify({"ok": False, "error": "admin login or valid sync token required"}), 403

    report_month = (payload.get("report_month") or request.form.get("report_month") or "").strip()
    mode = (payload.get("mode") or request.form.get("mode") or "replace").strip().lower()
    url = (
        payload.get("url")
        or request.form.get("url")
        or os.getenv("XWEATHER_GOOGLE_DRIVE_FILE_URL")
        or os.getenv("XWEATHER_GOOGLE_SHEET_CSV_URL")
        or DEFAULT_GOOGLE_DRIVE_FILE_URL
    )

    if not report_month:
        return jsonify({"ok": False, "error": "report_month required (YYYY-MM)"}), 400
    if mode not in ("append", "replace"):
        return jsonify({"ok": False, "error": "mode must be append or replace"}), 400

    try:
        raw, content_type, source_url = _download_google_workbook(url)
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": f"Google Drive file tidak bisa diakses server: {e}. Pastikan file dapat diakses publik atau gunakan token sync dari server."
        }), 502

    rows_data = []
    meta = {}
    if raw[:2] == b"PK" or "spreadsheet" in content_type or "excel" in content_type:
        rows_data, meta, meta_status = _rows_from_workbook(
            io.BytesIO(raw),
            report_month=report_month,
            require_month_sheet=True,
        )
        if meta_status != 200:
            return jsonify(meta), meta_status
    else:
        try:
            text = raw.decode("utf-8-sig")
        except Exception:
            text = raw.decode("latin-1")

        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return jsonify({"ok": False, "error": "Google Drive file has no readable Excel/CSV header"}), 400
        rows_data = list(reader)
        meta = {
            "selected_sheet": None,
            "required_sheet": _sheet_name_from_month(report_month),
        }

    init_db()
    result, status = _import_monthly_rows(rows_data, report_month, mode)
    result["source"] = "google_drive"
    result["source_url"] = source_url
    result["required_sheet"] = _sheet_name_from_month(report_month)
    result["selected_sheet"] = meta.get("selected_sheet")
    return jsonify(result), status


# =========================================================
# Expert Statistical Analysis (tab #3)
# GET /api/xweather/monthly-report/expert?month=YYYY-MM
# =========================================================
@xweather_report_bp.route("/api/xweather/monthly-report/expert", methods=["GET"])
def api_monthly_report_expert():
    month = _month_arg()
    if not month:
        return jsonify({"ok": False, "error": "missing month (YYYY-MM)"}), 400

    init_db()
    db = get_db()
    with db.cursor() as cur:
        cur.execute(f"""
            SELECT severity, {_event_time_expr()} AS analysis_event_time,
                   duration_min, strength_ka, event_type
            FROM monthly_lightning_alerts
            WHERE {_event_month_filter_expr()}
        """, _event_month_filter_params(month))
        rows = cur.fetchall()

    total = len(rows)
    if total == 0:
        return jsonify({
            "ok": True,
            "month": month,
            "metrics": [
                {"metric": "Total events", "value": 0},
                {"metric": "Period (UTC)", "value": f"{month}"},
            ]
        })

    sev_counts = Counter([(r["severity"] or "").strip().lower() for r in rows])
    alarm = sev_counts.get("alarm", 0)
    warning = sev_counts.get("warning", 0)
    info = sev_counts.get("info", 0)

    durations = []
    for r in rows:
        v = r.get("duration_min")
        if v is None:
            continue
        try:
            durations.append(float(v))
        except Exception:
            pass
    durations.sort()

    dmin = durations[0] if durations else None
    dmax = durations[-1] if durations else None
    dmed = _percentile(durations, 50) if durations else None
    dmean = (sum(durations) / len(durations)) if durations else None
    q25 = _percentile(durations, 25) if durations else None
    q75 = _percentile(durations, 75) if durations else None
    iqr = (q75 - q25) if (q25 is not None and q75 is not None) else None
    p90 = _percentile(durations, 90) if durations else None

    hours = []
    for r in rows:
        h = _parse_hour_utc(r.get("analysis_event_time"))
        if h is not None:
            hours.append(h)
    hour_counts = Counter(hours)
    most_hour = hour_counts.most_common(1)[0][0] if hour_counts else None

    busiest_window = None
    busiest_count = None
    if hour_counts:
        counts = [hour_counts.get(h, 0) for h in range(24)]
        best_h, best_sum = 0, -1
        for h in range(24):
            s3 = counts[h] + counts[(h + 1) % 24] + counts[(h + 2) % 24]
            if s3 > best_sum:
                best_sum = s3
                best_h = h
        busiest_window = f"{best_h:02d}:00–{(best_h + 3) % 24:02d}:00"
        busiest_count = best_sum

    types = []
    cg_minus = 0
    for r in rows:
        t = (r.get("event_type") or "").strip()
        if not t:
            continue
        tn = t.lower().replace(" ", "")
        types.append(t)
        if "cg-" in tn:
            cg_minus += 1
    dominant_type = Counter(types).most_common(1)[0][0] if types else None
    cg_minus_share = (cg_minus / total * 100.0) if total else 0.0

    strengths = []
    for r in rows:
        v = r.get("strength_ka")
        if v is None:
            continue
        try:
            strengths.append(abs(float(v)))
        except Exception:
            pass

    avg_abs_ka = (sum(strengths) / len(strengths)) if strengths else None
    min_abs_ka = min(strengths) if strengths else None
    max_abs_ka = max(strengths) if strengths else None

    metrics = [
        {"metric": "Total events", "value": total},
        {"metric": "Period (UTC)", "value": month},
        {"metric": "Fixed ring logic", "value": "Info=20 km; Warning=10 km; Alarm=4 km"},
        {"metric": "Alarm share", "value": f"{alarm} ({(alarm/total*100):.1f}%)"},
        {"metric": "Warning share", "value": f"{warning} ({(warning/total*100):.1f}%)"},
        {"metric": "Info share", "value": f"{info} ({(info/total*100):.1f}%)"},
    ]

    if durations:
        metrics += [
            {"metric": "Duration min / median / max (min)", "value": f"{dmin:.0f} / {dmed:.0f} / {dmax:.0f}"},
            {"metric": "Duration mean (min)", "value": f"{dmean:.2f}"},
            {"metric": "Duration IQR (Q75–Q25) (min)", "value": f"{iqr:.2f}" if iqr is not None else ""},
            {"metric": "Duration P90 (min)", "value": f"{p90:.2f}" if p90 is not None else ""},
        ]

    if most_hour is not None:
        metrics.append({"metric": "Most frequent hour (UTC)", "value": f"{most_hour}"})
    if busiest_window is not None:
        metrics.append({"metric": "Busiest 3-hour window (UTC)", "value": f"{busiest_window} ({busiest_count} events)"})

    if dominant_type:
        metrics.append({"metric": "Dominant type", "value": dominant_type})
    metrics.append({"metric": "CG- share", "value": f"{cg_minus_share:.1f}%"})

    if min_abs_ka is not None:
        metrics.append({"metric": "Minimum |peak current| (kA)", "value": f"{min_abs_ka:.2f}"})
    if avg_abs_ka is not None:
        metrics.append({"metric": "Average |peak current| (kA)", "value": f"{avg_abs_ka:.2f}"})
    if max_abs_ka is not None:
        metrics.append({"metric": "Maximum |peak current| (kA)", "value": f"{max_abs_ka:.2f}"})

    return jsonify({"ok": True, "month": month, "metrics": metrics})

@xweather_report_bp.route("/xweather/monthly-report.xlsx", methods=["GET"])
def monthly_report_xlsx():
    month = _month_arg()
    if not month:
        return "missing month (YYYY-MM)", 400

    init_db()
    db = get_db()
    with db.cursor() as cur:
        # 1) data report
        cur.execute(f"""
            SELECT severity, asset_name, extent_km, first_event_time, active_time,
                   last_event_time, clear_time, duration_min, strength_ka, event_type
            FROM monthly_lightning_alerts
            WHERE {_event_month_filter_expr()}
            ORDER BY {_event_sort_expr()} DESC NULLS LAST, id DESC
        """, _event_month_filter_params(month))
        report_rows = cur.fetchall()

        # 2) data expert (ambil data mentah yang dibutuhkan)
        cur.execute(f"""
            SELECT severity, {_event_time_expr()} AS analysis_event_time,
                   duration_min, strength_ka, event_type
            FROM monthly_lightning_alerts
            WHERE {_event_month_filter_expr()}
        """, _event_month_filter_params(month))
        expert_rows = cur.fetchall()

    # ===== Build Expert Metrics (mengikuti logic expert yang sudah ada) =====
    total = len(expert_rows)

    def _percentile(sorted_vals, p):
        if not sorted_vals:
            return None
        if len(sorted_vals) == 1:
            return float(sorted_vals[0])
        k = (len(sorted_vals) - 1) * (p / 100.0)
        f = int(k)
        c = min(f + 1, len(sorted_vals) - 1)
        if f == c:
            return float(sorted_vals[f])
        return float(sorted_vals[f] + (sorted_vals[c] - sorted_vals[f]) * (k - f))

    def _parse_hour_utc(s: str):
        if not s:
            return None
        s = str(s).strip()
        fmts = [
            "%d.%m.%Y %H:%M",
            "%d.%m.%Y %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%SZ",
            "%Y-%m-%dT%H:%M:%S",
        ]
        for f in fmts:
            try:
                return datetime.strptime(s, f).hour
            except Exception:
                pass
        m = re.search(r"\b([01]?\d|2[0-3])[:.]\d{2}\b", s)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return None
        return None

    sev_counts = Counter([(r["severity"] or "").strip().lower() for r in expert_rows])
    alarm = sev_counts.get("alarm", 0)
    warning = sev_counts.get("warning", 0)
    info = sev_counts.get("info", 0)

    durations = []
    for r in expert_rows:
        v = r.get("duration_min")
        if v is None:
            continue
        try:
            durations.append(float(v))
        except Exception:
            pass
    durations.sort()

    dmin = durations[0] if durations else None
    dmax = durations[-1] if durations else None
    dmed = _percentile(durations, 50) if durations else None
    dmean = (sum(durations) / len(durations)) if durations else None
    q25 = _percentile(durations, 25) if durations else None
    q75 = _percentile(durations, 75) if durations else None
    iqr = (q75 - q25) if (q25 is not None and q75 is not None) else None
    p90 = _percentile(durations, 90) if durations else None

    hours = []
    for r in expert_rows:
        h = _parse_hour_utc(r.get("analysis_event_time"))
        if h is not None:
            hours.append(h)
    hour_counts = Counter(hours)
    most_hour = hour_counts.most_common(1)[0][0] if hour_counts else None

    busiest_window = None
    busiest_count = None
    if hour_counts:
        counts = [hour_counts.get(h, 0) for h in range(24)]
        best_h, best_sum = 0, -1
        for h in range(24):
            s3 = counts[h] + counts[(h + 1) % 24] + counts[(h + 2) % 24]
            if s3 > best_sum:
                best_sum = s3
                best_h = h
        busiest_window = f"{best_h:02d}:00–{(best_h + 3) % 24:02d}:00"
        busiest_count = best_sum

    types = []
    cg_minus = 0
    for r in expert_rows:
        t = (r.get("event_type") or "").strip()
        if not t:
            continue
        tn = t.lower().replace(" ", "")
        types.append(t)
        if "cg-" in tn:
            cg_minus += 1
    dominant_type = Counter(types).most_common(1)[0][0] if types else None
    cg_minus_share = (cg_minus / total * 100.0) if total else 0.0

    strengths = []
    for r in expert_rows:
        v = r.get("strength_ka")
        if v is None:
            continue
        try:
            strengths.append(abs(float(v)))
        except Exception:
            pass
    min_abs_ka = min(strengths) if strengths else None
    max_abs_ka = max(strengths) if strengths else None
    avg_abs_ka = (sum(strengths) / len(strengths)) if strengths else None

    expert_metrics = [
        ("Total events", total),
        ("Period (UTC)", month),
        ("Fixed ring logic", "Info=20 km; Warning=10 km; Alarm=4 km"),
        ("Alarm share", f"{alarm} ({(alarm/total*100):.1f}%)" if total else "0 (0.0%)"),
        ("Warning share", f"{warning} ({(warning/total*100):.1f}%)" if total else "0 (0.0%)"),
        ("Info share", f"{info} ({(info/total*100):.1f}%)" if total else "0 (0.0%)"),
    ]

    if durations:
        expert_metrics += [
            ("Duration min / median / max (min)", f"{dmin:.0f} / {dmed:.0f} / {dmax:.0f}"),
            ("Duration mean (min)", f"{dmean:.2f}"),
            ("Duration IQR (Q75–Q25) (min)", f"{iqr:.2f}" if iqr is not None else ""),
            ("Duration P90 (min)", f"{p90:.2f}" if p90 is not None else ""),
        ]

    if most_hour is not None:
        expert_metrics.append(("Most frequent hour (UTC)", str(most_hour)))
    if busiest_window is not None:
        expert_metrics.append(("Busiest 3-hour window (UTC)", f"{busiest_window} ({busiest_count} events)"))

    if dominant_type:
        expert_metrics.append(("Dominant type", dominant_type))
    expert_metrics.append(("CG- share", f"{cg_minus_share:.1f}%"))

    if min_abs_ka is not None:
        expert_metrics.append(("Minimum |peak current| (kA)", f"{min_abs_ka:.2f}"))
    if avg_abs_ka is not None:
        expert_metrics.append(("Average |peak current| (kA)", f"{avg_abs_ka:.2f}"))
    if max_abs_ka is not None:
        expert_metrics.append(("Maximum |peak current| (kA)", f"{max_abs_ka:.2f}"))

    # ===== Build Excel workbook with 2 sheets =====
    wb = Workbook()

    # Sheet 1: Report
    ws1 = wb.active
    ws1.title = "Monthly Report"
    headers = [
        "Severity","Asset name","Extent (km)","First event time","Active time",
        "Last event time","Clear time","Duration (min)","Strength (kA)","Type"
    ]
    ws1.append(headers)
    for r in report_rows:
        ws1.append([
            r.get("severity"),
            r.get("asset_name"),
            r.get("extent_km"),
            r.get("first_event_time"),
            r.get("active_time"),
            r.get("last_event_time"),
            r.get("clear_time"),
            r.get("duration_min"),
            r.get("strength_ka"),
            r.get("event_type"),
        ])

    # autosize (simple)
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 2: Expert
    ws2 = wb.create_sheet("Expert Statistical Analysis")
    ws2.append(["Metric", "Value"])
    for k, v in expert_metrics:
        ws2.append([k, str(v)])

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 42

    # write to bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = make_response(bio.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="Xweather_Monthly_Report_{month}.xlsx"'
    return resp

@xweather_report_bp.route("/api/xweather/golf-locations", methods=["GET"])
def api_list_golf_locations():
    init_db()
    db = get_db()
    with db.cursor() as cur:
        cur.execute("SELECT * FROM golf_locations ORDER BY name ASC")
        rows = cur.fetchall()
    return jsonify({"ok": True, "rows": rows})

@xweather_report_bp.route("/api/xweather/golf-locations", methods=["POST"])
def api_create_golf_location():
    init_db()
    payload = request.get_json(force=True)

    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            INSERT INTO golf_locations(
                name, latitude, longitude,
                alarm_km, warning_km, info_km,
                alarm_color, warning_color, info_color,
        
                rule2_enabled, rule2_min_strikes, rule2_window_sec,
        
                created_at
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING *
        """, (
            payload["name"],
            payload["latitude"],
            payload["longitude"],
            payload.get("alarm_km", 4),
            payload.get("warning_km", 10),
            payload.get("info_km", 20),
            payload.get("alarm_color", "#EF4444"),
            payload.get("warning_color", "#F97316"),
            payload.get("info_color", "#FBBF24"),
        
            bool(payload.get("rule2_enabled", True)),
            int(payload.get("rule2_min_strikes", 1)),
            int(payload.get("rule2_window_sec", 60)),
        
            _now_iso()
        ))
        row = cur.fetchone()
    db.commit()
    return jsonify({"ok": True, "row": row})

@xweather_report_bp.route("/api/xweather/golf-locations/<int:id>", methods=["PUT"])
def api_update_golf_location(id):
    init_db()
    payload = request.get_json(force=True)
    db = get_db()

    with db.cursor() as cur:
        cur.execute("""
            UPDATE golf_locations SET
                name=%s,
                latitude=%s,
                longitude=%s,
                alarm_km=%s,
                warning_km=%s,
                info_km=%s,
                alarm_color=%s,
                warning_color=%s,
                info_color=%s,
        
                rule2_enabled=%s,
                rule2_min_strikes=%s,
                rule2_window_sec=%s
        
            WHERE id=%s
            RETURNING *
        """, (
            payload["name"],
            payload["latitude"],
            payload["longitude"],
            payload["alarm_km"],
            payload["warning_km"],
            payload["info_km"],
            payload["alarm_color"],
            payload["warning_color"],
            payload["info_color"],
        
            bool(payload.get("rule2_enabled", True)),
            int(payload.get("rule2_min_strikes", 1)),
            int(payload.get("rule2_window_sec", 60)),
        
            id
        ))
        row = cur.fetchone()

    db.commit()
    return jsonify({"ok": True, "row": row})

@xweather_report_bp.route("/api/xweather/golf-locations/<int:id>", methods=["DELETE"])
def api_delete_golf_location(id):
    init_db()
    db = get_db()
    with db.cursor() as cur:
        cur.execute("DELETE FROM golf_locations WHERE id=%s RETURNING id",(id,))
        row = cur.fetchone()
    db.commit()
    return jsonify({"ok": True})

@xweather_report_bp.route("/xweather/golfareamap")
@login_required
@role_required("admin")
def xweather_golf_area_map_demo():
    init_db()
    return render_template("project/xweather_golf_map_admin.html")

try:
    import websocket  # websocket-client
except Exception:
    websocket = None

# --- simple pubsub for SSE ---
_blitz_clients = set()
_blitz_lock = threading.Lock()
_blitz_thread_started = False
_blitz_buffer = deque(maxlen=20000)   # simpan strike terbaru
_blitz_buf_lock = threading.Lock()

def _sse_format(data: dict) -> str:
    return f"data: {json.dumps(data, ensure_ascii=False)}\n\n"

def _decode_obfuscated_payload(payload_bytes: bytes) -> bytes:
    """
    NOTE: Blitzortung WS payload is obfuscated (LZW-like).
    This decode function is taken from public reverse engineering references.
    Adjust if protocol changes.
    """
    e = {}
    d = list(payload_bytes.decode(errors="ignore"))
    if not d:
        return b""
    c = d[0]
    f = c
    g = [c]
    h = 256
    o = h
    for i in range(1, len(d)):
        a = ord(d[i])
        a = d[i] if h > a else e[a] if e.get(a) else f + c
        g.append(a)
        c = a[0]
        e[o] = f + c
        o += 1
        f = a
    return "".join(g).encode()

def _norm_polarity(v):
    """
    Normalize polarity to "+", "-", or None.
    Supports numeric (+1/-1), string ("+", "-", "CG+", "CG-", "pos/neg").
    """
    if v is None:
        return None

    # numeric
    if isinstance(v, (int, float)):
        if v > 0:
            return "+"
        if v < 0:
            return "-"
        return None

    s = str(v).strip().lower()
    if not s:
        return None

    # direct tokens
    if s in {"+", "pos", "positive", "plus"}:
        return "+"
    if s in {"-", "neg", "negative", "minus"}:
        return "-"

    # embedded in type strings
    if "cg+" in s or "pos" in s:
        return "+"
    if "cg-" in s or "neg" in s:
        return "-"

    return None

def _start_blitz_ws_thread():
    global _blitz_thread_started
    if _blitz_thread_started:
        return
    _blitz_thread_started = True

    def run():
        if websocket is None:
            print("[Blitz] websocket-client not installed.")
            return

        WS_URL = "wss://ws1.blitzortung.org"  # public reference
        # handshake message reference: {"a":111}
        HELLO = {"a": 111}

        while True:
            try:
                ws = websocket.create_connection(WS_URL, timeout=15)
                ws.send(json.dumps(HELLO))
                print("[Blitz] connected")

                while True:
                    msg = ws.recv()
                    if msg is None:
                        raise RuntimeError("WS closed")

                    # msg can be str or bytes
                    if isinstance(msg, str):
                        raw = msg.encode()
                    else:
                        raw = msg

                    # try decode obfuscation -> json
                    decoded = _decode_obfuscated_payload(raw)
                    if not decoded:
                        continue

                    try:
                        j = json.loads(decoded.decode("utf-8", errors="ignore"))
                    except Exception:
                        continue

                    # Normalize to what frontend needs
                    # Depending on message type, fields may differ; keep defensive:
                    lat = j.get("lat") or j.get("latitude")
                    lon = j.get("lon") or j.get("lng") or j.get("longitude")
                    ts  = j.get("time") or j.get("timestamp") or j.get("t")

                    if lat is None or lon is None:
                        continue

                    raw_type = j.get("type") or j.get("s") or "strike"

                    # --- added: polarity extraction (defensive) ---
                    raw_pol = (
                        j.get("pol") or
                        j.get("polarity") or
                        j.get("sgn") or
                        j.get("sign") or
                        j.get("p")
                    )
                    pol = _norm_polarity(raw_pol) or _norm_polarity(raw_type)
                    # -------------------------------------------

                    out = {
                        "lat": float(lat),
                        "lon": float(lon),
                        "time": ts,
                        "type": raw_type,
                        "polarity": pol,  # "+", "-", or None
                    }
                    out["_t_rcv"] = time.time()  # epoch seconds saat diterima server

                    with _blitz_buf_lock:
                        _blitz_buffer.append(out)

                    # broadcast to all SSE clients
                    with _blitz_lock:
                        dead = []
                        for q in _blitz_clients:
                            try:
                                q.put_nowait(out)
                            except Exception:
                                dead.append(q)
                        for q in dead:
                            _blitz_clients.discard(q)

            except Exception as e:
                print("[Blitz] reconnecting after error:", e)
                try:
                    ws.close()
                except Exception:
                    pass
                time.sleep(3)

    t = threading.Thread(target=run, daemon=True)
    t.start()


@xweather_report_bp.route("/api/blitzortung/sse")
def api_blitzortung_sse():
    """
    Server-Sent Events endpoint:
    browser will connect and receive stream of lightning points.
    """
    _start_blitz_ws_thread()

    q = queue.Queue(maxsize=2000)
    with _blitz_lock:
        _blitz_clients.add(q)

    def gen():
        # initial ping
        yield _sse_format({"ok": True, "msg": "connected"})
        try:
            while True:
                item = q.get()
                yield _sse_format(item)
        except GeneratorExit:
            pass
        finally:
            with _blitz_lock:
                _blitz_clients.discard(q)

    return Response(gen(), mimetype="text/event-stream")


@xweather_report_bp.route("/api/blitzortung/last")
def api_blitzortung_last():
    """
    Return strikes received in last N seconds (default 60).
    """
    _start_blitz_ws_thread()

    try:
        seconds = int(request.args.get("seconds", "60"))
    except Exception:
        seconds = 60
    seconds = max(1, min(seconds, 3600))  # 1..3600

    cutoff = time.time() - seconds
    items = []
    with _blitz_buf_lock:
        # ambil yang terbaru saja
        for it in reversed(_blitz_buffer):
            if float(it.get("_t_rcv", 0)) < cutoff:
                break
            items.append(it)

    items.reverse()
    return jsonify({"ok": True, "seconds": seconds, "count": len(items), "items": items})
